""" 
说明：
    运行此脚本，可以将原选课名单的学生均匀地分成[助教人数]组
    该脚本只应运行一次。在之后有新的学生选课后，应运行updategroup.py，以免已分组学生的分组发生变化。
"""

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

OUTPUT_FILENAME = '成绩考核登记表-分组.xlsx'
INPUT_FILE_SUFFIX = '成绩考核登记表.xlsx'

# 找到未分组的成绩考核登记表
def find_filename_with_suffix(s : str):
    curr_dir_filelist = [f for f in os.listdir('.') if os.path.isfile(f)]
    for f in curr_dir_filelist:
        if f.endswith(INPUT_FILE_SUFFIX):
            return f
    return ''
curr_dir_filelist = [f for f in os.listdir('.') if os.path.isfile(f)]
filename = ''
for f in curr_dir_filelist:
    if f.endswith(INPUT_FILE_SUFFIX):
        filename = f
if filename == '':
    print('表格未找到')
    exit(0)

# 打开表格，实施分组
grade_wb = load_workbook(filename)
main_ws = grade_wb['All']
if len(grade_wb.worksheets):
    other_sheets = grade_wb.worksheets.copy()
    other_sheets.remove(grade_wb['All'])
    for s in other_sheets:
        grade_wb.remove(s)

NUM_GROUP = 5
NUM_ROW_HEADER = 7

def copy_sheet_header(main_ws:Worksheet, dst_ws:Worksheet):
    # determine max column
    #dst_ws.merge_cells('A1:M7')
    dst_ws.merged_cells.ranges = main_ws.merged_cells.ranges
    for r in main_ws.iter_rows(1, NUM_ROW_HEADER):
        for c in r:
            #dst_ws.cell(c.row, c.column, c.value)
            dst_ws.copy_cell(c.row, c.column, c)

sub_group_sheets = []
# sub_group_sheet_end_idx = []
sub_group_sheet_stu_num = []

for group_idx in range(NUM_GROUP):
    sheet_name = chr(ord('A') + group_idx) + '组'
    sub_ws = grade_wb.create_sheet(sheet_name)
    copy_sheet_header(main_ws, sub_ws)
    sub_group_sheets.append(sub_ws)
    #sub_group_sheet_end_idx.append(NUM_ROW_HEADER + 1)
    sub_group_sheet_stu_num.append(0)

for r in main_ws.iter_rows(NUM_ROW_HEADER + 1, 300):
    # fetch stu id in this class
    if r[0].value is None:
        continue
    remains = (int(r[0].value)-1) % NUM_GROUP
    for c in r:
        dst_ws = sub_group_sheets[remains]
        next_idx = NUM_ROW_HEADER + 1 + sub_group_sheet_stu_num[remains]
        dst_ws.copy_cell_to_row(next_idx, c.column, c)
    sub_group_sheet_stu_num[remains] += 1


grade_wb.save(OUTPUT_FILENAME)
print('分组后的表格已保存。各组的人数为：')
for i in range(NUM_GROUP):
    print(sub_group_sheet_stu_num[i])