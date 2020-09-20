""" 
说明：
    运行此脚本，可以将新选课名单中找出增加的新同学，并加入到已分组的列表中。之前已分组的同学的组号不会改变。
    在之后有新的学生选课后，应运行updategroup.py，以免已分组学生的分组发生变化。
输入：

"""

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os

import utils
from utils import OUTPUT_FILENAME, INPUT_FILE_SUFFIX, NUM_GROUP

existing_stu_names = set() # will be dirty
new_stu_names = set()

new_stu_names_ordered = []

# existing names
grouped_wb = load_workbook(OUTPUT_FILENAME)
a = grouped_wb.sheetnames
all_student_sheet = grouped_wb['All'] if 'All' in grouped_wb.sheetnames else grouped_wb['完整名单']
for r in all_student_sheet.iter_rows(8, 400):
    if r[1].value is None:
        continue
    existing_stu_names.add(r[1].value)

num_existing_student = len(existing_stu_names)

input_filename = utils.find_filename_with_suffix(INPUT_FILE_SUFFIX)
if(input_filename == ''):
    print('未找到新登记表')
    exit(0)
new_wb = load_workbook(input_filename)
new_student_sheet = new_wb['All'] if 'All' in new_wb.sheetnames else new_wb['完整名单']

for r in new_student_sheet.iter_rows(8, 400):
    if r[1].value is None:
        continue
    new_stu_names.add(r[1].value)
    new_stu_names_ordered.append(r[1].value)

# check number of grouped students

num_group_student = []

for i in range(NUM_GROUP):
    group_str = chr(ord('A') + i) + '组'
    curr_sheet = grouped_wb[group_str]

    num_group_student.append(0)

    for r in curr_sheet.iter_rows(8, 400):
        if r[1].value is None:
            continue
        num_group_student[i] += 1

# find quitted students
quitted_stu = []
for idx, stuno in enumerate(existing_stu_names):
    if stuno not in new_stu_names:
        print('学生' + stuno + '已退课')
        quitted_stu.append(stuno)
        
        # remove from all
        for ridx, r in enumerate(all_student_sheet.iter_rows(8, 400)):
            if r[1].value == stuno:
                all_student_sheet.delete_rows(ridx)
                print('已从总表中删除')
                num_existing_student -= 1
                break
        # remove from the group
        for gidx in range(NUM_GROUP):
            group_str = chr(ord('A') + gidx) + '组'
            curr_sheet = grouped_wb[group_str]
            for ridx, r in enumerate(curr_sheet.iter_rows(8, 400)):
                if r[1].value == stuno:
                    curr_sheet.delete_rows(ridx)
                    print('已从分组中删除')
                    num_group_student[gidx] -= 1
                    break

# find new added students
new_added_stu = []
for idx, stuno in enumerate(new_stu_names_ordered):
    if stuno not in existing_stu_names:
        print('发现新学生' + stuno)
        new_added_stu.append((stuno, idx))


for new_stu in new_added_stu:
    # 复制新学生到分组后的All表
    # 新学生ridx
    new_stu_ridx = 0
    for ridx, r in enumerate(new_student_sheet.iter_rows(8, 400)):
        if r[1].value == new_stu[0]:
            new_stu_ridx = ridx
    src_row = new_stu_ridx + 8 # an existing row
    dst_row = num_existing_student + 8 # a blank row
    for r in new_student_sheet.iter_rows(src_row, src_row):
        for cell in r:
            all_student_sheet.copy_cell_to_row(dst_row, cell.column, cell)
    num_existing_student += 1

    # 分配新学生到组
    # 若有某组比别人都少，则分配至该组
    min_idx = 0
    min_group_num = num_group_student[0]
    for i in range(1,NUM_GROUP):
        if num_group_student[i] < min_group_num:
            min_group_num, min_idx = num_group_student[i], i
    group_str = chr(ord('A') + min_idx) + '组'
    print('分配学生' + new_stu[0] + '到' + group_str)
    curr_sheet = grouped_wb[group_str]
    dst_row_group = num_group_student[min_idx] + 8
    for r in new_student_sheet.iter_rows(src_row, src_row):
        for cell in r:
            curr_sheet.copy_cell_to_row(dst_row_group, cell.column, cell)
    num_group_student[min_idx] += 1

grouped_wb.save(OUTPUT_FILENAME)